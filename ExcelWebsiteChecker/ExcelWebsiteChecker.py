from openpyxl import load_workbook
import requests
from threading import Thread

INPUT_SPREADSHEET_NAME = "list.xlsx"
OUTPUT_SPREADSHEET_NAME = "list_updated.xlsx"
LOG_ERRORS_TO_OUTPUT = False

def checkWebsite(threadName, website):
    httpsWebAddress = "https://" + website.replace("https://","").replace("http://", "").replace("http//:","").replace("http:","") # Hack there are some malformed websites in the spreadsheet
    #httpWebAddress = "http://" + website
    try:
        #print(threadName+"Attempting to hit website: " + httpsWebAddress)
        ret = requests.head(httpsWebAddress, timeout=5)
        print(threadName+"Website "+httpsWebAddress+" returned status code: " + str(ret.status_code))
        return "ALIVE"
    except requests.exceptions.SSLError as e:
        if LOG_ERRORS_TO_OUTPUT:
            print(threadName+"SSL Error occured for website: " + httpsWebAddress+ " Assuming this means the site is Dead as a real buisness probably would have a valid HTTPS cert.")#. Trying http...")
        return "DEAD"
        #try:
        #    print("Attempting to hit website: " + httpWebAddress)
        #    ret = requests.head(httpWebAddress)
        #    print("Website returned status code: " + str(ret.status_code))
        #except requests.exceptions.NewConnectionError as e:
        #    print("Website " + website + " is possibly dead ", e)
        #    return False
    except requests.ConnectionError as e:
        if LOG_ERRORS_TO_OUTPUT:
            print(threadName+"Website " + website + " is possibly dead.\n", e)
        return "DEAD"
    except requests.exceptions.ReadTimeout as e:
        if LOG_ERRORS_TO_OUTPUT:
            print(threadName+"Website "+website+" timed out.", e)
        return "TIMEOUT"

def checkSheet(threadName, sheet):
    print(threadName+"##################################################")
    print(threadName+"Opening sheet: " + sheet.title)
    print(threadName+"##################################################")
   
    firstRow = tuple(sheet.rows)[0]
    webAddressHeaderIndex = [item.value for item in firstRow].index("Web Address")
    for index, row in enumerate(list(sheet.rows)[1:-1], start=2): # [1:-1] to skip header row and final row (because it will be empty), start=2 to have index start at 2 for enumeration
        website = row[webAddressHeaderIndex].value
        if website is None: # HACK - openpyxl was giving me issues where some rows at the end of a sheet, with nothing in them would be enumerated.
            print("WANRING FOUND A WEBSITE WITH 'NONE' VALUE Logging output: Sheet Name: "+sheet.title+" Current Row: "+str(row[0].row)+" Row Count: "+str(len(list(sheet.rows))))
            return
        #print("Row: " + str(index) + " Value: " + website)
        status = checkWebsite(threadName, website)
        print(threadName+"Website "+website+" is: "+status)
        # update row "Website Status" column
        row[webAddressHeaderIndex+1].value = status # TODO: This currently assumes the website status column is directly to the right of the Web Address column

workbook = load_workbook('list.xlsx')
print(workbook.sheetnames)
threads = []
for sheet in workbook:
    sheetName = sheet.title
    # One thread per sheet, with a unique name for logging
    process = Thread(target=checkSheet, args=["##"+sheetName+"## ", sheet])
    process.start()
    threads.append(process)

# Wait for threads to complete
for process in threads:
    process.join()

print("All sheets have been checked. Saving updates to new Excel spreadsheet...")
workbook.save("list_updated.xlsx")
print("New spreadsheet saved. list_updated.xlsx")